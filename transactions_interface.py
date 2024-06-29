
# Contact Manager using python tkinter and openpyxl
from tkinter import Tk, Entry, Label, Button
import openpyxl
from datetime import datetime

import os
print (os.getcwd()) 

# In this case we have to manually create *.xlsx file in project directory
# REASON: the file created from python file context manager will have empty metadata
# with no cell formatting which will raise 'Badzipfile' error
# with open("contact_book.xlsx", "a"):
#     pass

# loading workbook
# mybook = openpyxl.load_workbook("/home/aashish/Documents/Python-Journey/Python-Projects/Contact_Manager/contact_book.xlsx")
# mybook = openpyxl.load_workbook(r"C:\Users\sluo\Desktop\JL\00_PythonWIP\2024.3.30_PyProjects_Demos\PyProjects-main-BEGINNER-FRIENDLY\Contact_Manager\contact_book.xlsx")
mybook = openpyxl.load_workbook(r".\New Sheet.xlsx")
sheet = mybook['Transactions']


def load_workbook(sheet, mybook):
    """Adding different field.get() if there doesn't exist any"""
    cellassign = sheet.cell(1, 1)
    if cellassign.value != "Investor":
        sheet['A1'] = "Investor"
        sheet['B1'] = "ID"
        sheet['C1'] = "Transaction Date"
        sheet['D1'] = "Ticker"
        sheet['E1'] = "Type"
        sheet['F1'] = "Shares"
        sheet['G1'] = "Cost Per Share"
        sheet['H1'] = "Transaction Total"
        mybook.save("New Sheet.xlsx")

def check_write_data(*args):
    """Checking if user input match with database"""
    row = sheet.max_row
    # for i in range(1, row + 1):
    #     name_obj = sheet.cell(i, 1)
    #     phone_obj = sheet.cell(i, 2)

    #     if investor_field.get() == "" or phone_field.get() == "":
    #         warning_text = Label(root, text="Please, Enter Valid Input!",
    #                              background="black", fg="red", font=("garuda", 11, "bold"))
    #         warning_text.grid(row=11, column=1)
    #         return False

    #     # Raising error if there exist any simialr contact
    #     # if (name_obj.value == investor_field.get()) or (phone_obj.value == phone_field.get()):
    #     #     warning_text = Label(root, text="Contact Already Exist\nChange your Name or Phone number and Try Again!",
    #     #                          background="black", fg="red", font=("garuda", 11, "bold"))
    #     #     warning_text.grid(row=11, column=1)
    #     #     return False

    # # Writing data if there doesn't exist any duplication
    # else:
    data = ((investor_field.get(), id_field.get(), tarnsaction_date_field.get(),
            ticker_field.get(), type_field.get(), int(shares_field.get()), int(cost_per_share_field.get()),
            int(transaction_total_field.get())))
    # for row in data:
        # sheet.append(row)
    sheet.append(data)
    mybook.save("New Sheet.xlsx")

    warning_text = Label(root, text="Transaction Successfully Saved",
                            background="black", fg="green", font=("garuda", 11, "bold"))
    warning_text.grid(row=14, column=1, ipadx=100, ipady=50)
    return True


# UI development
# Customizing root window
root = Tk()
root.geometry("588x433")
root.minsize(588, 433)
root.maxsize(688, 533)
root.configure(background="black")
root.title("Contact Manager")

# Adding header
root_text = Label(root, text="Transaction Request", background="black",
                  fg="white", font=("garuda", 18, "bold"))
root_text.grid(column=1)

# Intializing Label
# Investor, ID, tarnsaction_date, Ticker, Type;
investor = Label(root, text="Investor", background="black", foreground="white")
id = Label(root, text="ID", background="black", foreground="white")
tarnsaction_date = Label(root, text="Tarnsaction Date", background="black", foreground="white")
ticker = Label(root, text="Ticker", background="black", foreground="white")
type = Label(root, text="Type", background="black", foreground="white")
shares = Label(root, text="Shares", background="black", foreground="white")
cost_per_share = Label(root, text="Cost Per Share", background="black", foreground="white")
transaction_total = Label(root, text="Transaction Total", background="black", foreground="white")

# Packing Label as Grid
investor.grid(row=5, column=0, ipadx=40)
id.grid(row=6, column=0)
tarnsaction_date.grid(row=7, column=0)
ticker.grid(row=8, column=0)
type.grid(row=9, column=0)
shares.grid(row=10, column=0)
cost_per_share.grid(row=11, column=0)
transaction_total.grid(row=12, column=0)

# Creating Entry Field
investor_field = Entry(root, background="black", foreground="white")
id_field = Entry(root, background="black", foreground="white")
tarnsaction_date_field = Entry(root, background="black", foreground="white")
ticker_field = Entry(root, background="black", foreground="white")
type_field = Entry(root, background="black", foreground="white")
shares_field = Entry(root, background="black", foreground="white")
cost_per_share_field = Entry(root, background="black", foreground="white")
transaction_total_field = Entry(root, background="black", foreground="white")

# Packing Entry field as grid
investor_field.grid(row=5, column=1, pady=3, ipadx=100)
id_field.grid(row=6, column=1, pady=3, ipadx=100)
tarnsaction_date_field.grid(row=7, column=1, pady=3, ipadx=100)
ticker_field.grid(row=8, column=1, pady=3, ipadx=100)
type_field.grid(row=9, column=1, pady=3, ipadx=100)
shares_field.grid(row=10, column=1, pady=3, ipadx=100)
cost_per_share_field.grid(row=11, column=1, pady=3, ipadx=100)
transaction_total_field.grid(row=12, column=1, pady=3, ipadx=100)

# Creating Button
# load_workbook('Transactions', mybook)
load_workbook(sheet, mybook)
submit_button = Button(root, text="Save", background="black",
                       foreground="white", command=check_write_data)
submit_button.grid(row=13, column=1, pady=15, ipadx=10)

# Creating Close Button
def close_app():
    root.destroy()

close_button = Button(root, text="Close", background="black",
                      foreground="white", command=close_app)
close_button.grid(row=13, column=2, pady=15, ipadx=10)

# Creating Reset Button
def reset_form():
    # Resetting the form
    investor_field.delete(0, 100)
    id_field.delete(0, 100)
    tarnsaction_date_field.delete(0, 100)
    ticker_field.delete(0, 100)
    type_field.delete(0, 100)
    shares_field.delete(0, 100)
    cost_per_share_field.delete(0, 100)
    transaction_total_field.delete(0, 100)

reset_button = Button(root, text="Reset", background="black",
                      foreground="white", command=reset_form)
reset_button.grid(row=13, column=0, pady=15, ipadx=10)

# Focusing to  next entry field if "Enter" is pressed/"<Return>"
# But first initializing input to name entry field
investor_field.bind("<Return>", func=investor_field.focus_set())
investor_field.bind("<Return>", lambda event: id_field.focus_set())
id_field.bind("<Return>", lambda event: tarnsaction_date_field.focus_set())
tarnsaction_date_field.bind("<Return>", lambda event: ticker_field.focus_set())
ticker_field.bind("<Return>", lambda event: type_field.focus_set())
type_field.bind("<Return>", lambda event: shares_field.focus_set())
shares_field.bind("<Return>", lambda event: cost_per_share_field.focus_set())
cost_per_share_field.bind("<Return>", lambda event: transaction_total_field.focus_set())
transaction_total_field.bind("<Return>", lambda event: submit_button.focus_set())

submit_button.bind("<Return>", func=check_write_data)
investor_field.focus_set()

root.mainloop()
